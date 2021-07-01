from openpyxl import Workbook, load_workbook
import re
import sys
import os

class WorkbookController:

    def initialise(self):
        self.workbook = load_workbook(filename=os.path.join(sys.path[0], 'spotify-plays.xlsx.xlsm'))
        self.sheet = self.workbook.active
        artist_link = []
        for row in range(2, self.sheet.max_row):
            if self.sheet.cell(row=row, column=1).value is None:
                break
            else:
                artist_link.append(self.sheet.cell(row=row, column=1).value)

        return artist_link

    def get_song_id(self, link):
        song_id = re.search("track/(.*?)\?si=", link).group(1)
        return song_id

    def get_artist_id(self, link):
        artist_id = re.search("artist/(.*?)\?si=", link).group(1)
        return artist_id


    def save_workbook(self):
        self.workbook.save(filename=os.path.join(sys.path[0], 'spotify-plays.xlsx.xlsm'))


